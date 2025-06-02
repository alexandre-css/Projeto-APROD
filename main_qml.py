from PySide6.QtCore import QObject, Signal as pyqtSignal, Property as pyqtProperty, Slot as pyqtSlot
from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox, QMainWindow, QListWidget
from PySide6.QtQml import QQmlApplicationEngine
from xls2xlsx import XLS2XLSX
import sys, os, shutil
import pandas as pd
import re
import json
import datetime
import math
import csv
import traceback
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

os.environ["QT_QUICK_CONTROLS_STYLE"] = "Material"

MESES_PT = {
    'Jan': 'jan', 'Feb': 'fev', 'Mar': 'mar', 'Apr': 'abr', 'May': 'mai', 'Jun': 'jun',
    'Jul': 'jul', 'Aug': 'ago', 'Sep': 'set', 'Oct': 'out', 'Nov': 'nov', 'Dec': 'dez'
}

class Backend(QObject):
    nomesChanged = pyqtSignal()
    valoresChanged = pyqtSignal()
    arquivosCarregadosChanged = pyqtSignal()
    tabelaPesosChanged = pyqtSignal()
    mesesAtivosChanged = pyqtSignal()
    tabelaSemanaChanged = pyqtSignal()
    rankingSemanaChanged = pyqtSignal()
    tabelaSemanaOrdenadaChanged = pyqtSignal()
    rankingModelChanged = pyqtSignal(list)
    rankingModelPopulated = pyqtSignal(list)
    sortColumnChanged = pyqtSignal(str)
    sortAscendingChanged = pyqtSignal(bool)
    mesAtivoChanged = pyqtSignal(str, bool)
    opacityChanged = pyqtSignal(str, float)
    kpisChanged = pyqtSignal()
    chartUpdateRequired = pyqtSignal()
    pageLoaded = pyqtSignal(str)
    pesosModelUpdate = pyqtSignal()
    filtroChanged = pyqtSignal(str)
    pesosModelChanged = pyqtSignal(list)
    pesosModelDataReady = pyqtSignal()
    exportSuccessSignal = pyqtSignal()
    opacityUpdateRequired = pyqtSignal(str, float)
    mesStatusUpdateRequired = pyqtSignal(str, bool)

    def __init__(self, mainwindow):
        super().__init__()
        self.mainwindow = mainwindow
        self.gerenciador_pesos = mainwindow.gerenciador_pesos
        self._nomes = []
        self._valores = []
        self._mesesAtivos = []
        self._arquivosCarregados = ""
        self._tabelaSemana = []
        self._rankingSemana = ""
        self._sortColumn = "usuario"
        self._sortAscending = True
        self._rankingData = []
        self._currentPage = "dashboard"
        self._filtroUsuario = ""

    def conectar_sinais_ui(self):
        self.sortColumnChanged.connect(lambda col: print(f"Coluna ordenação mudou: {col}"))
        self.sortAscendingChanged.connect(lambda asc: print(f"Ordem mudou: {asc}"))
        self.rankingModelChanged.connect(lambda dados: print(f"Ranking atualizado: {len(dados)} itens"))
        self.mesAtivoChanged.connect(lambda mes, ativo: print(f"Mês {mes} {'ativado' if ativo else 'desativado'}"))

    def _filtrar_arquivos_por_meses_ativos(self, arquivos):
        print(f"Meses ativos para filtro: {self._mesesAtivos}")
        print(f"Arquivos antes do filtro: {arquivos}")
        if not self._mesesAtivos:
            print("Nenhum mês ativo, retornando todos os arquivos")
            return arquivos  

        result = []
        print(f"Meses ativos: {self._mesesAtivos}")
        
        for arq in arquivos:
            file_path = os.path.join("dados_mensais", arq)
            periodo = self.mainwindow.extrair_mes_ano_do_arquivo(file_path)
            meses_ativos_lower = [m.lower() for m in self._mesesAtivos]
            
            print(f"Arquivo: {arq}, Período: {periodo}")
            print(f"Período em minúsculas: {periodo.lower() if periodo else None}")
            print(f"Está nos meses ativos? {periodo.lower() in meses_ativos_lower if periodo else False}")
            
            if periodo and periodo.lower() in meses_ativos_lower:
                result.append(arq)
        
        print(f"Filtragem: {len(arquivos)} -> {len(result)} arquivos")
        return result
    
    def _coletar_tipos_brutos(self):
        pasta = "dados_mensais"
        arquivos = [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")] if os.path.exists(pasta) else []
        
        arquivos = self._filtrar_arquivos_por_meses_ativos(arquivos)
        
        tipos = set()
        for arq in arquivos:
            file_path = os.path.join(pasta, arq)
            try:
                df = pd.read_excel(file_path, skiprows=1, usecols="G")
                if len(df.columns) > 0:
                    df.columns = ["Agendamento"]
                    tipos_array = df["Agendamento"].dropna().astype(str).str.strip()
                    tipos_array = tipos_array.apply(lambda x: re.sub(r'\s*\(.*?\)', '', x).strip())
                    tipos_unicos = set(tipos_array.tolist())
                    tipos.update(tipos_unicos)
                    print(f"Arquivo {arq}: {len(tipos_unicos)} tipos únicos encontrados")
            except Exception as e:
                print(f"Erro coletando tipos brutos de {arq}: {str(e)}")
        
        resultado = sorted(tipos)
        print(f"Total de tipos únicos coletados: {len(resultado)}")
        print(f"Tipos encontrados: {resultado}")
        return resultado

    def normalizar_chave(self, dia):
        if hasattr(self, '_normalizacao_cache'):
            return self._normalizacao_cache.get(dia, dia.lower())
        
        self._normalizacao_cache = {
            "Segunda": "segunda",
            "Terça": "terca",
            "Quarta": "quarta",
            "Quinta": "quinta",
            "Sexta": "sexta",
            "Sábado": "sabado",
            "Domingo": "domingo",
        }
        return self._normalizacao_cache.get(dia, dia.lower())

    @pyqtSlot()
    def atualizar_pesos_model_ui(self):
        try:
            self.atualizar_tabela_pesos()
            self.pesosModelChanged.emit(self.pesosModel)
        except Exception as e:
            print(f"Erro ao atualizar pesos model UI: {str(e)}")

    @pyqtSlot()
    def popular_pesos_model(self):
        try:
            pesos_data = []
            for i, tipo in enumerate(self._nomes):
                peso = self._valores[i] if i < len(self._valores) else 1.0
                pesos_data.append({"tipo": tipo, "peso": peso})
            self.pesosModelChanged.emit(self.pesosModel)
            self.pesosModelDataReady.emit()
        except Exception as e:
            print(f"Erro ao popular modelo de pesos: {str(e)}")

    @pyqtProperty(list, notify=pesosModelChanged)
    def pesosModel(self):
        try:
            return [
                {"tipo": nome, "peso": valor}
                for nome, valor in zip(self._nomes, self._valores)
            ]
        except Exception as e:
            print(f"Erro ao gerar pesosModel: {str(e)}")
            return []

    @pyqtSlot()
    def atualizar_pesos_model_direto(self):
        try:
            self.atualizar_tabela_pesos()
            self.pesosModelChanged.emit(self.pesosModel)
        except Exception as e:
            print(f"Erro ao atualizar modelo direto: {str(e)}")

    @pyqtSlot(str)
    def filtrarTabelaSemana(self, filtro):
        try:
            self._filtroUsuario = filtro
            filtered_data = []
            
            if not self._tabelaSemana:
                return
                
            for item in self._tabelaSemana:
                if not filtro or filtro.lower() in item.get('usuario', '').lower():
                    filtered_data.append(item)
            
            self._tabelaSemana = filtered_data
            self.tabelaSemanaChanged.emit()
        except Exception as e:
            print(f"Erro ao filtrar tabela semana: {str(e)}")
            import traceback
            traceback.print_exc()
    
    @pyqtSlot(str)
    def exportarTabelaSemana(self, formato):
        try:
            if not self._tabelaSemana:
                print("Nenhum dado para exportar")
                return
            
            pasta_docs = os.path.join(os.path.dirname(__file__), "exports", "tabela_semana")
            try:
                os.makedirs(pasta_docs, exist_ok=True)
                print(f"Pasta criada com sucesso: {pasta_docs}")
            except Exception as e:
                print(f"Erro ao criar pasta: {e}")
                return
            print(f"Pasta existe? {os.path.exists(pasta_docs)}")
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if formato.lower() == "csv":
                nome_arquivo = f"tabela_semana_{timestamp}.csv"
                caminho_arquivo = os.path.join(pasta_docs, nome_arquivo)
                print(f"Tentando criar CSV em: {caminho_arquivo}")
                
                with open(caminho_arquivo, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile, delimiter=';')
                    
                    meses_ref = " - ".join(self._mesesAtivos) if self._mesesAtivos else "Todos os períodos"
                    writer.writerow(['RELATÓRIO DE PRODUTIVIDADE SEMANAL'])
                    writer.writerow([f'Período: {meses_ref}'])
                    writer.writerow([f'Gerado em: {datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")}'])
                    writer.writerow([])
                    
                    writer.writerow(['Usuário', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo', 'Total'])
                    
                    for item in self._tabelaSemana:
                        total = sum(item.get(dia, 0) for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo'])
                        row = [
                            item.get('usuario', ''),
                            f"{item.get('segunda', 0):.1f}",
                            f"{item.get('terca', 0):.1f}",
                            f"{item.get('quarta', 0):.1f}",
                            f"{item.get('quinta', 0):.1f}",
                            f"{item.get('sexta', 0):.1f}",
                            f"{item.get('sabado', 0):.1f}",
                            f"{item.get('domingo', 0):.1f}",
                            f"{total:.1f}"
                        ]
                        writer.writerow(row)
                print(f"CSV criado com sucesso: {caminho_arquivo}")
                self.exportSuccessSignal.emit()
            
            elif formato.lower() == "xlsx":
                nome_arquivo = f"tabela_semana_{timestamp}.xlsx"
                caminho_arquivo = os.path.join(pasta_docs, nome_arquivo)
                print(f"Tentando criar XLSX em: {caminho_arquivo}")
                
                data = []
                for item in self._tabelaSemana:
                    total = sum(item.get(dia, 0) for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo'])
                    row = {
                        'Usuário': item.get('usuario', ''),
                        'Segunda': item.get('segunda', 0),
                        'Terça': item.get('terca', 0),
                        'Quarta': item.get('quarta', 0),
                        'Quinta': item.get('quinta', 0),
                        'Sexta': item.get('sexta', 0),
                        'Sábado': item.get('sabado', 0),
                        'Domingo': item.get('domingo', 0),
                        'Total': total
                    }
                    data.append(row)
                
                df = pd.DataFrame(data)
                
                with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Produtividade Semanal', index=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Produtividade Semanal']
                    
                    meses_ref = " - ".join(self._mesesAtivos) if self._mesesAtivos else "Todos os períodos"
                    worksheet.insert_rows(1)
                    worksheet.cell(row=1, column=1).value = f"Período: {meses_ref}"
                    
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    
                    info_fill = PatternFill(start_color='e8f4fd', end_color='e8f4fd', fill_type='solid')
                    info_font = Font(bold=True, size=11, color='3cb3e6')
                    worksheet.cell(row=1, column=1).fill = info_fill
                    worksheet.cell(row=1, column=1).font = info_font
                    worksheet.merge_cells('A1:I1')
                    
                    header_fill = PatternFill(start_color='3cb3e6', end_color='3cb3e6', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True, size=12)
                    data_font = Font(size=10)
                    center_align = Alignment(horizontal='center', vertical='center')
                    border = Border(
                        left=Side(style='thin', color='dee2e6'),
                        right=Side(style='thin', color='dee2e6'),
                        top=Side(style='thin', color='dee2e6'),
                        bottom=Side(style='thin', color='dee2e6')
                    )
                    
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=2, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                        cell.border = border
                    
                    for row in range(3, len(df) + 3):
                        for col in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = data_font
                            cell.alignment = center_align
                            cell.border = border
                            
                            if col % 2 == 0:
                                cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
                    
                    for col_num, col in enumerate(worksheet.columns, 1):
                        try:
                            max_length = 0
                            for cell in col:
                                if hasattr(cell, 'value') and cell.value is not None:
                                    max_length = max(max_length, len(str(cell.value)))
                            
                            col_letter = chr(64 + col_num)
                            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 20)
                        except Exception as e:
                            print(f"Erro ao ajustar largura da coluna {col_num}: {e}")
                            continue
                
                print(f"XLSX criado com sucesso: {caminho_arquivo}")
                self.exportSuccessSignal.emit()
            
            elif formato.lower() == "pdf":
                nome_arquivo = f"tabela_semana_{timestamp}.pdf"
                caminho_arquivo = os.path.join(pasta_docs, nome_arquivo)
                print(f"Tentando criar PDF em: {caminho_arquivo}")
                
                doc = SimpleDocTemplate(caminho_arquivo, pagesize=A4)
                elements = []
                
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=20,
                    spaceAfter=20,
                    alignment=1,
                    textColor=colors.HexColor('#3cb3e6'),
                    fontName='Helvetica-Bold'
                )
                
                subtitle_style = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Normal'],
                    fontSize=12,
                    spaceAfter=20,
                    alignment=1,
                    textColor=colors.HexColor('#666666')
                )
                
                title = Paragraph("RELATÓRIO DE PRODUTIVIDADE SEMANAL", title_style)
                elements.append(title)
                
                meses_ref = " - ".join(self._mesesAtivos) if self._mesesAtivos else "Todos os períodos"
                periodo_subtitle = Paragraph(f"Período: {meses_ref}", subtitle_style)
                elements.append(periodo_subtitle)
                
                data_atual = datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")
                data_subtitle = Paragraph(f"Gerado em {data_atual}", subtitle_style)
                elements.append(data_subtitle)
                elements.append(Spacer(1, 20))
                
                data = [['Usuário', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo', 'Total']]
                
                for item in self._tabelaSemana:
                    total = sum(item.get(dia, 0) for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo'])
                    usuario = item.get('usuario', '')
                    if len(usuario) > 15:
                        usuario = usuario.replace('.', '.\n')
                    row = [
                        usuario,
                        f"{item.get('segunda', 0):.1f}",
                        f"{item.get('terca', 0):.1f}",
                        f"{item.get('quarta', 0):.1f}",
                        f"{item.get('quinta', 0):.1f}",
                        f"{item.get('sexta', 0):.1f}",
                        f"{item.get('sabado', 0):.1f}",
                        f"{item.get('domingo', 0):.1f}",
                        f"{total:.1f}"
                    ]
                    data.append(row)
                
                table = Table(data, colWidths=[100, 45, 45, 45, 45, 45, 45, 45, 50])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3cb3e6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('TOPPADDING', (0, 0), (-1, 0), 10),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#dee2e6')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (0, -1), 4),
                    ('RIGHTPADDING', (0, 0), (0, -1), 4),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    ('BACKGROUND', (-1, 1), (-1, -1), colors.HexColor('#e3f2fd')),
                    ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold')
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 30))
                
                footer_style = ParagraphStyle(
                    'Footer',
                    parent=styles['Normal'],
                    fontSize=10,
                    alignment=1,
                    textColor=colors.HexColor('#666666')
                )
                footer = Paragraph("Relatório gerado pelo Sistema de Análise de Produtividade", footer_style)
                elements.append(footer)
                
                doc.build(elements)
                print(f"PDF criado com sucesso: {caminho_arquivo}")
                self.exportSuccessSignal.emit()
            
            print(f"Arquivo {formato.upper()} exportado com sucesso para: {caminho_arquivo}")
            
        except Exception as e:
            print(f"Erro ao exportar {formato.upper()}: {str(e)}")
            import traceback
            traceback.print_exc()

    @pyqtSlot(str, bool)
    def ordenarTabelaSemana(self, coluna, ascendente):
        try:
            self._sortColumn = coluna
            self._sortAscending = ascendente
            
            if not self._tabelaSemana:
                return
            
            if coluna == "Total":
                self._tabelaSemana.sort(
                    key=lambda x: sum(x.get(dia, 0) for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo']),
                    reverse=not ascendente
                )
            elif coluna == "usuario":
                self._tabelaSemana.sort(key=lambda x: x.get('usuario', '').lower(), reverse=not ascendente)
            else:
                self._tabelaSemana.sort(key=lambda x: x.get(coluna, 0), reverse=not ascendente)
            
            self.tabelaSemanaChanged.emit()
            self.sortColumnChanged.emit(coluna)
            self.sortAscendingChanged.emit(ascendente)
        except Exception as e:
            print(f"Erro ao ordenar tabela: {str(e)}")

    @pyqtSlot(str, float)
    def set_mes_opacity(self, mes, opacity):
        try:
            self.opacityUpdateRequired.emit(mes, opacity)
        except Exception as e:
            print(f"Erro ao definir opacidade: {str(e)}")

    @pyqtSlot(str)
    def toggle_mes_ativo_ui(self, mes):
        try:
            self.toggleMesAtivo(mes)
        except Exception as e:
            print(f"Erro ao alternar mês ativo: {str(e)}")

    @pyqtSlot(str)
    def handle_page_loaded(self, page):
        try:
            self._currentPage = page
            self.pageLoaded.emit(page)
            
            if page == "dashboard":
                self.atualizar_arquivos_carregados()
                self.atualizar_grafico()
                self.atualizar_kpis()
                self.force_chart_update()
            elif page == "semana":
                self.gerarTabelaSemana()
                self.atualizar_ranking_model()
            elif page == "pesos":
                self.atualizar_tabela_pesos()
                self.popular_pesos_model()
        except Exception as e:
            print(f"Erro ao carregar página {page}: {str(e)}")

    @pyqtSlot()
    def force_chart_update(self):
        try:
            self.chartUpdateRequired.emit()
        except Exception as e:
            print(f"Erro ao forçar atualização do gráfico: {str(e)}")

    @pyqtSlot()
    def atualizar_kpis_ui(self):
        try:
            self.atualizar_kpis()
        except Exception as e:
            print(f"Erro ao atualizar KPIs UI: {str(e)}")

    @pyqtSlot()
    def restore_pesos_complete(self):
        try:
            tipos_atuais = self._coletar_tipos_brutos()
            self.gerenciador_pesos.restaurar_pesos_padrao(tipos_atuais)
            
            self._nomes = tipos_atuais
            self._valores = [1 for _ in tipos_atuais]
            
            self.nomesChanged.emit()
            self.valoresChanged.emit()
            self.tabelaPesosChanged.emit()
            self.pesosModelChanged.emit(self.pesosModel)
            print("Pesos restaurados com sucesso")
        except Exception as e:
            print(f"Erro ao restaurar pesos: {str(e)}")

    @pyqtSlot()
    def gerarTabelaSemana(self):
        try:
            pasta = "dados_mensais"
            arquivos = [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")] if os.path.exists(pasta) else []
            
            arquivos = self._filtrar_arquivos_por_meses_ativos(arquivos)
            
            dfs = []
            for arq in arquivos:
                file_path = os.path.join(pasta, arq)
                df = self.mainwindow.carregar_planilha(file_path)
                if df is not None and not df.empty:
                    dfs.append(df)
            
            if not dfs:
                self._tabelaSemana = []
                self.tabelaSemanaChanged.emit()
                return
            
            df_total = pd.concat(dfs, ignore_index=True)
            df_total = df_total.loc[:, ~df_total.columns.duplicated()]
            
            print(f"Total de linhas processadas: {len(df_total)}")
            print(f"Colunas disponíveis: {df_total.columns.tolist()}")
            
            usuarios_dados = {}
            
            if "Usuário" in df_total.columns and "Data criação" in df_total.columns:
                df_total["Data criação"] = pd.to_datetime(df_total["Data criação"], errors='coerce', dayfirst=True)
                
                datas_ainda_invalidas = df_total["Data criação"].isna()
                if datas_ainda_invalidas.any():
                    df_total.loc[datas_ainda_invalidas, "Data criação"] = pd.to_datetime(
                        df_total.loc[datas_ainda_invalidas, "Data criação"], 
                        errors='coerce', 
                        dayfirst=False
                    )
                
                datas_ainda_invalidas = df_total["Data criação"].isna()
                if datas_ainda_invalidas.any():
                    df_total.loc[datas_ainda_invalidas, "Data criação"] = pd.to_datetime(
                        df_total.loc[datas_ainda_invalidas, "Data criação"], 
                        errors='coerce', 
                        format='%d/%m/%Y'
                    )
                
                datas_invalidas = df_total["Data criação"].isna().sum()
                print(f"Datas inválidas encontradas: {datas_invalidas}")
                print(f"Datas válidas após conversão: {df_total['Data criação'].notna().sum()}")
                
                df_total["DiaSemana"] = df_total["Data criação"].dt.day_name()
                
                mapeamento_dias = {
                    'Monday': 'segunda', 'Tuesday': 'terca', 'Wednesday': 'quarta',
                    'Thursday': 'quinta', 'Friday': 'sexta', 'Saturday': 'sabado', 'Sunday': 'domingo'
                }
                df_total["DiaSemana"] = df_total["DiaSemana"].map(mapeamento_dias)
                
                print(f"Registros com usuário válido: {df_total['Usuário'].notna().sum()}")
                print(f"Registros com dia da semana válido: {df_total['DiaSemana'].notna().sum()}")
                
                for _, row in df_total.iterrows():
                    usuario = row["Usuário"]
                    dia = row["DiaSemana"]
                    peso = float(row["peso"]) if pd.notna(row["peso"]) and "peso" in df_total.columns else 1.0
                    
                    if pd.isna(usuario) or usuario == "":
                        continue
                    
                    if pd.isna(dia):
                        dia = "segunda"
                        
                    if usuario not in usuarios_dados:
                        usuarios_dados[usuario] = {
                            'usuario': usuario,
                            'segunda': 0, 'terca': 0, 'quarta': 0, 'quinta': 0,
                            'sexta': 0, 'sabado': 0, 'domingo': 0
                        }
                    
                    if dia in usuarios_dados[usuario]:
                        usuarios_dados[usuario][dia] += peso
            
            self._tabelaSemana = list(usuarios_dados.values())
            total_minutas = sum(sum(user.get(dia, 0) for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo']) for user in usuarios_dados.values())
            print(f"TabelaSemana gerada com {len(self._tabelaSemana)} usuários únicos")
            print(f"Total de minutas contabilizadas: {total_minutas}")
            self.tabelaSemanaChanged.emit()
            
        except Exception as e:
            print(f"Erro ao gerar tabela semana: {str(e)}")

    @pyqtSlot()
    def atualizar_ranking_model(self):
        try:
            if not self._tabelaSemana:
                self._rankingData = []
                self.rankingModelChanged.emit([])
                return
            
            dias_totais = {}
            for item in self._tabelaSemana:
                for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo']:
                    if dia not in dias_totais:
                        dias_totais[dia] = 0
                    dias_totais[dia] += item.get(dia, 0)
            
            ranking_sorted = sorted(dias_totais.items(), key=lambda x: x[1], reverse=True)
            
            dias_nomes = {
                'segunda': 'Segunda-feira', 'terca': 'Terça-feira', 'quarta': 'Quarta-feira',
                'quinta': 'Quinta-feira', 'sexta': 'Sexta-feira', 'sabado': 'Sábado', 'domingo': 'Domingo'
            }
            
            ranking_data = []
            for i, (dia, total) in enumerate(ranking_sorted):
                nome_dia = dias_nomes.get(dia, dia.capitalize())
                ranking_data.append({
                    "rankText": f"{i+1}º {nome_dia}: {total:.1f}",
                    "dia": dia,
                    "total": total,
                    "posicao": i+1
                })
            
            self._rankingData = ranking_data
            self.rankingModelChanged.emit(ranking_data)
            
        except Exception as e:
            print(f"Erro ao atualizar ranking: {str(e)}")

    @pyqtSlot()
    def popular_ranking_model_completo(self):
        try:
            self.atualizar_ranking_model()
        except Exception as e:
            print(f"Erro ao popular ranking completo: {str(e)}")

    @pyqtSlot()
    def popular_ranking_model_direto(self):
        try:
            self.atualizar_ranking_model()
        except Exception as e:
            print(f"Erro ao popular ranking direto: {str(e)}")

    @pyqtSlot(str)
    def atualizar_sort_table(self, column):
        try:
            if self._sortColumn == column:
                self._sortAscending = not self._sortAscending
            else:
                self._sortColumn = column
                self._sortAscending = True
            
            self.ordenarTabelaSemana(column, self._sortAscending)
            
        except Exception as e:
            print(f"Erro ao atualizar ordenação: {str(e)}")

    @pyqtSlot(str)
    def setFiltroUsuario(self, filtro):
        self._filtroUsuario = filtro
        self.filtroChanged.emit(filtro)
        self.filtrarTabelaSemana(filtro)

    @pyqtSlot()
    def exportar_csv_ui(self):
        try:
            self.exportarTabelaSemanaCSV()
        except Exception as e:
            print(f"Erro ao exportar CSV UI: {str(e)}")

    @pyqtSlot(dict, result=float)
    def calculate_row_total(self, item_data):
        try:
            total = 0
            for dia in ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo']:
                total += item_data.get(dia, 0)
            return total
        except Exception as e:
            print(f"Erro ao calcular total da linha: {str(e)}")
            return 0

    @pyqtSlot(dict, str, result=float)
    def calculate_day_value(self, user_data, day_key):
        try:
            return user_data.get(day_key, 0)
        except Exception as e:
            print(f"Erro ao calcular valor do dia: {str(e)}")
            return 0

    @pyqtSlot(int)
    def decrementar_peso(self, index):
        try:
            if 0 <= index < len(self._valores):
                novo_peso = max(0.1, self._valores[index] - 0.1)
                self._valores[index] = round(novo_peso, 1)
                
                if index < len(self._nomes):
                    tipo = self._nomes[index]
                    self.mainwindow.gerenciador_pesos.pesos[tipo.strip().upper()] = novo_peso
                
                self.pesosModelChanged.emit(self.pesosModel)
                self.valoresChanged.emit()
        except Exception as e:
            print(f"Erro ao decrementar peso: {str(e)}")

    @pyqtSlot(int)
    def incrementar_peso(self, index):
        try:
            if 0 <= index < len(self._valores):
                novo_peso = min(10.0, self._valores[index] + 0.1)
                self._valores[index] = round(novo_peso, 1)
                
                if index < len(self._nomes):
                    tipo = self._nomes[index]
                    self.mainwindow.gerenciador_pesos.pesos[tipo.strip().upper()] = novo_peso
                
                self.pesosModelChanged.emit(self.pesosModel)
                self.valoresChanged.emit()
        except Exception as e:
            print(f"Erro ao incrementar peso: {str(e)}")

    @pyqtSlot(int, str)
    def validar_peso_input(self, index, texto):
        try:
            if 0 <= index < len(self._valores):
                try:
                    novo_peso = float(texto)
                    novo_peso = max(0.1, min(10.0, novo_peso))
                    self._valores[index] = round(novo_peso, 1)
                    
                    if index < len(self._nomes):
                        tipo = self._nomes[index]
                        self.mainwindow.gerenciador_pesos.pesos[tipo.strip().upper()] = novo_peso
                    
                    self.pesosModelChanged.emit(self.pesosModel)
                    self.valoresChanged.emit()
                except ValueError:
                    print(f"Valor inválido: {texto}")
        except Exception as e:
            print(f"Erro ao validar peso: {str(e)}")

    @pyqtSlot(dict, str, result=str)
    def get_day_value_formatted(self, user_data, day_key):
        try:
            value = user_data.get(day_key, 0)
            return f"{value:.1f}" if value > 0 else "0"
        except Exception as e:
            print(f"Erro ao formatar valor do dia: {str(e)}")
            return "0"

    @pyqtSlot(dict, result=str)
    def get_row_total_formatted(self, item_data):
        try:
            total = self.calculate_row_total(item_data)
            return f"{total:.1f}" if total > 0 else "0"
        except Exception as e:
            print(f"Erro ao formatar total da linha: {str(e)}")
            return "0"

    @pyqtSlot(dict, str, result=float)
    def get_day_value(self, user_data, day_key):
        return self.calculate_day_value(user_data, day_key)

    @pyqtSlot(dict, result=float)
    def get_row_total(self, item_data):
        return self.calculate_row_total(item_data)
    
    @pyqtSlot(str)
    def atualizar_estado_mes(self, mes):
        try:
            is_active = mes in self._mesesAtivos
            self.mesStatusUpdateRequired.emit(mes, is_active)
        except Exception as e:
            print(f"Erro ao atualizar estado do mês: {str(e)}")

    @pyqtSlot(str, float)
    def aplicar_opacity_mes(self, mes, opacity):
        try:
            self.opacityUpdateRequired.emit(mes, opacity)
        except Exception as e:
            print(f"Erro ao aplicar opacidade: {str(e)}")

    @pyqtSlot()
    def inicializar_pesos_model(self):
        try:
            self.atualizar_tabela_pesos()
            self.popular_pesos_model()
        except Exception as e:
            print(f"Erro ao inicializar pesos model: {str(e)}")

    @pyqtSlot()
    def recarregar_dados_completo(self):
        try:
            self.atualizar_arquivos_carregados()
            self.atualizar_grafico()
            self.atualizar_kpis()
            self.atualizar_tabela_pesos()
        except Exception as e:
            print(f"Erro ao recarregar dados: {str(e)}")

    @pyqtSlot()
    def gerar_tabela_semana_completa(self):
        try:
            self.gerarTabelaSemana()
            self.atualizar_ranking_model()
        except Exception as e:
            print(f"Erro ao gerar tabela semana completa: {str(e)}")

    @pyqtSlot()
    def refresh_chart_loader(self):
        try:
            self.chartUpdateRequired.emit()
        except Exception as e:
            print(f"Erro ao atualizar loader do gráfico: {str(e)}")

    @pyqtSlot(str)
    def toggleMesAtivo(self, mes):
        QApplication.processEvents()
        
        if mes in self._mesesAtivos:
            self._mesesAtivos.remove(mes)
        else:
            self._mesesAtivos.append(mes)
        
        self.mesesAtivosChanged.emit()
        print(f"Atualizando dados após toggle de {mes}, meses ativos: {self._mesesAtivos}")
        
        # IMPORTANTE: Primeiro notificar mudanças dos valores para QML reconfigurar o gráfico
        self.valoresChanged.emit()
        self.nomesChanged.emit()
        
        # Depois atualizar os dados em ordem específica
        QApplication.processEvents()  # Garante processamento dos sinais antes de continuar
        self.atualizar_kpis()
        QApplication.processEvents()  # Pequena pausa para processamento
        self.atualizar_tabela_pesos()
        QApplication.processEvents()  # Pequena pausa para processamento
        self.atualizar_grafico()  # Deixa o gráfico por último
        
        # Notificar os KPIs após todos os dados estarem prontos
        if hasattr(self.mainwindow, "kpis_qml") and self.mainwindow.kpis_qml is not None:
            self.mainwindow.kpis_qml.kpisChanged.emit()

    @pyqtSlot()
    def atualizar_kpis(self):
        try:
            pasta = "dados_mensais"
            arquivos = [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")] if os.path.exists(pasta) else []
            arquivos = self._filtrar_arquivos_por_meses_ativos(arquivos)
            
            if not arquivos:
                if hasattr(self.mainwindow, "kpis_qml") and self.mainwindow.kpis_qml:
                    self.mainwindow.kpis_qml.atualizar_kpis("0", "0", "--", "--")
                return
            
            dfs = []
            for arq in arquivos:
                file_path = os.path.join(pasta, arq)
                df = self.mainwindow.carregar_planilha(file_path)
                if not df.empty:
                    dfs.append(df)
            
            if not dfs:
                if hasattr(self.mainwindow, "kpis_qml") and self.mainwindow.kpis_qml:
                    self.mainwindow.kpis_qml.atualizar_kpis("0", "0", "--", "--")
                return
            
            df_total = pd.concat(dfs, ignore_index=True)
            
            total_minutas = len(df_total)
            
            dia_mais_produtivo = "--"
            if "Data criação" in df_total.columns and "peso" in df_total.columns:
                df_total["Data criação"] = pd.to_datetime(df_total["Data criação"], errors='coerce', dayfirst=True)
                
                datas_ainda_invalidas = df_total["Data criação"].isna()
                if datas_ainda_invalidas.any():
                    df_total.loc[datas_ainda_invalidas, "Data criação"] = pd.to_datetime(
                        df_total.loc[datas_ainda_invalidas, "Data criação"], 
                        errors='coerce', 
                        dayfirst=False
                    )
                
                df_total = df_total[df_total["Data criação"].notna()]
                
                if not df_total.empty:
                    df_total["DiaSemana"] = df_total["Data criação"].dt.day_name()
                    
                    mapeamento_dias = {
                        'Monday': 'Segunda', 'Tuesday': 'Terça', 'Wednesday': 'Quarta',
                        'Thursday': 'Quinta', 'Friday': 'Sexta', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
                    }
                    df_total["DiaSemana"] = df_total["DiaSemana"].map(mapeamento_dias)
                    
                    produtividade_dias = df_total.groupby("DiaSemana")["peso"].sum()
                    if not produtividade_dias.empty:
                        dia_mais_produtivo_nome = produtividade_dias.idxmax()
                        valor_dia_mais_produtivo = produtividade_dias.max()
                        if valor_dia_mais_produtivo >= 1000:
                            dia_mais_produtivo = f"{dia_mais_produtivo_nome}\n({valor_dia_mais_produtivo:.0f})"
                        else:
                            dia_mais_produtivo = f"{dia_mais_produtivo_nome} ({valor_dia_mais_produtivo:.1f})"
            
            top3_text = "--"
            if "Usuário" in df_total.columns and "peso" in df_total.columns:
                top_usuarios = df_total.groupby("Usuário")["peso"].sum().sort_values(ascending=False).head(3)
                top3_lines = []
                for i, (user, peso) in enumerate(top_usuarios.items(), 1):
                    top3_lines.append(f"{i}º {user}: {int(peso)}")
                top3_text = " | ".join(top3_lines)
            
            if hasattr(self.mainwindow, "kpis_qml") and self.mainwindow.kpis_qml:
                self.mainwindow.kpis_qml.atualizar_kpis(str(total_minutas), "0", dia_mais_produtivo, top3_text)
            
            self.kpisChanged.emit()
            
        except Exception as e:
            print(f"Erro ao calcular KPIs: {str(e)}")
            if hasattr(self.mainwindow, "kpis_qml") and self.mainwindow.kpis_qml:
                self.mainwindow.kpis_qml.atualizar_kpis("0", "0", "--", "--")
    
    @pyqtSlot()
    def atualizar_arquivos_carregados(self):
        pasta = "dados_mensais"
        
        if not os.path.exists(pasta):
            self._arquivosCarregados = "Nenhum arquivo carregado"
            self._mesesAtivos = []
        else:
            arquivos = sorted([arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")])
            periodos = []
            for arq in arquivos:
                file_path = os.path.join(pasta, arq)
                periodo = self.mainwindow.extrair_mes_ano_do_arquivo(file_path)
                if periodo and "/" in periodo:
                    periodos.append(periodo)
            
            self._arquivosCarregados = " - ".join(periodos) if periodos else "Nenhum arquivo carregado"
            
            if not hasattr(self, "_mesesAtivos") or not self._mesesAtivos:
                self._mesesAtivos = periodos.copy()
                    
        self.arquivosCarregadosChanged.emit()
        self.mesesAtivosChanged.emit()
    
    @pyqtSlot()
    def atualizar_grafico(self):
        pasta = "dados_mensais"
        arquivos = [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")] if os.path.exists(pasta) else []
        
        arquivos = self._filtrar_arquivos_por_meses_ativos(arquivos)
        
        dfs = []
        for arq in arquivos:
            file_path = os.path.join(pasta, arq)
            df = self.mainwindow.carregar_planilha(file_path)
            if "Usuário" in df.columns:
                df["Usuário"] = df["Usuário"].astype(str).str.strip().str.upper()
                df = df[~df["Usuário"].isin(["ANGELOBRASIL", "SECAUTOLOC"])]
            df = df.loc[:, ~df.columns.duplicated()]
            if not df.empty:
                dfs.append(df)
    
        if not dfs:
            self._nomes = ["Sem dados disponíveis"]
            self._valores = [0]
            self.nomesChanged.emit()
            self.valoresChanged.emit()
            return
    
        df_total = pd.concat(dfs, ignore_index=True)
        df_total = df_total.loc[:, ~df.columns.duplicated()]
        
        if "Usuário" in df_total.columns and "peso" in df_total.columns:
            df_total = df_total[df_total["Usuário"].notna() & (df_total["Usuário"] != "")]
            df_total = df_total[~df_total["Usuário"].isin(["ANGELOBRASIL", "SECAUTOLOC"])]
            produtividade = df_total.groupby("Usuário")["peso"].sum()
            produtividade = produtividade[produtividade.index.notna()]
            produtividade = produtividade.sort_values(ascending=False)
            
            self._nomes = list(produtividade.index)[:20]
            self._valores = [max(1.0, float(v)) for v in produtividade.values[:20]]
        else:
            self._nomes = []
            self._valores = []
    
        self.nomesChanged.emit()
        self.valoresChanged.emit()
    
    @pyqtSlot()
    def importar_arquivo_excel(self):
        self.mainwindow.importar_arquivo_excel()
        self.atualizar_arquivos_carregados()
        self.atualizar_kpis()
        self.atualizar_grafico()
        self.atualizar_tabela_pesos()

    @pyqtSlot(str, result=bool)
    def mes_ativo_status(self, mes):
        try:
            return mes in self._mesesAtivos
        except Exception as e:
            print(f"Erro ao verificar status do mês: {str(e)}")
            return False

    @pyqtProperty(float, notify=chartUpdateRequired)
    def maxAxisValue(self):
        try:
            if not self._valores or len(self._valores) == 0:
                return 20.0
            valores_validos = [float(v) for v in self._valores if v is not None]
            if not valores_validos:
                return 20.0
            max_val = max(valores_validos)
            return max(20.0, math.ceil(max_val * 1.1))
        except Exception as e:
            print(f"Erro ao calcular maxAxisValue: {str(e)}")
            return 20.0

    @pyqtProperty(str, notify=sortColumnChanged)
    def sortColumn(self):
        return self._sortColumn

    @pyqtProperty(bool, notify=sortAscendingChanged)  
    def sortAscending(self):
        return self._sortAscending

    @pyqtProperty(list, notify=rankingModelChanged)
    def rankingData(self):
        return self._rankingData

    @pyqtProperty('QStringList', notify=mesesAtivosChanged)
    def mesesAtivos(self):
        return self._mesesAtivos

    @pyqtProperty(list, notify=arquivosCarregadosChanged)
    def mesesDisponiveis(self):
        pasta = "dados_mensais"
        periodos = []
        
        if os.path.exists(pasta):
            arquivos = [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")]
            print(f"⚠️ Arquivos encontrados na pasta: {len(arquivos)}")
            
            for arquivo in arquivos:
                file_path = os.path.join(pasta, arquivo)
                periodo = self.mainwindow.extrair_mes_ano_do_arquivo(file_path)
                if periodo and "/" in periodo:
                    periodos.append(periodo)
                    print(f"Adicionando período: {periodo}")
        

        result = sorted(list(set(periodos)), key=self.mainwindow.chave_mes_ano)
        print(f"⚠️ Total de períodos para exibição: {len(result)}, meses = {result}")
        return result

    @pyqtProperty(list, notify=nomesChanged)
    def nomes(self):
        return self._nomes

    @pyqtProperty(list, notify=valoresChanged)
    def valores(self):
        return self._valores

    @pyqtProperty(list, notify=tabelaPesosChanged)
    def tabela_pesos(self):
        return [
            {'tipo': nome, 'peso': valor}
            for nome, valor in zip(self._nomes, self._valores)
        ]

    @pyqtProperty(str, notify=arquivosCarregadosChanged)
    def arquivosCarregados(self):
        return self._arquivosCarregados

    @pyqtProperty(list, notify=tabelaSemanaChanged)
    def tabelaSemana(self):
        return self._tabelaSemana

    @pyqtProperty(str, notify=rankingSemanaChanged)
    def rankingSemana(self):
        return self._rankingSemana
    
    @pyqtProperty(str, notify=pageLoaded)
    def currentPage(self):
        return self._currentPage
    
    @pyqtProperty(str)
    def filtroUsuario(self):
        return self._filtroUsuario

    @pyqtProperty(bool)
    def temDadosDisponiveis(self):
        return bool(self._tabelaSemana and len(self._tabelaSemana) > 0)

    @pyqtProperty(int)
    def totalItensTabela(self):
        return len(self._tabelaSemana) if self._tabelaSemana else 0

    @pyqtProperty(bool)
    def graficoTemDados(self):
        return bool(self._valores and len(self._valores) > 0 and any(v > 0 for v in self._valores))

    @pyqtProperty('QVariant', notify=kpisChanged)
    def kpis(self):
        try:
            if hasattr(self.mainwindow, 'kpis_qml') and self.mainwindow.kpis_qml:
                return {
                    'minutas': self.mainwindow.kpis_qml.minutas,
                    'dia_produtivo': self.mainwindow.kpis_qml.dia,
                    'top3': self.mainwindow.kpis_qml.top3
                }
            return {'minutas': '0', 'dia_produtivo': '--', 'top3': '--'}
        except Exception as e:
            return {'minutas': '0', 'dia_produtivo': '--', 'top3': '--'}


    #& Aqui começam os métodos de atribuição de pesos.

    @pyqtSlot()
    def atualizar_tabela_pesos(self):
        try:
            tipos_atuais = self._coletar_tipos_brutos()
            
            self._nomes = tipos_atuais
            self._valores = []
            
            for tipo in tipos_atuais:
                peso = self.mainwindow.gerenciador_pesos.obter_peso(tipo)
                self._valores.append(peso)
            
            self.nomesChanged.emit()
            self.valoresChanged.emit()
            self.tabelaPesosChanged.emit()
            self.pesosModelChanged.emit(self.pesosModel)
        except Exception as e:
            print(f"Erro ao atualizar tabela de pesos: {str(e)}")

    @pyqtSlot()
    def abrirPastaExports(self):
        try:
            import subprocess
            import platform
            
            pasta_exports = os.path.join(os.path.dirname(__file__), "exports", "tabela_semana")
            
            if platform.system() == "Windows":
                subprocess.run(["explorer", pasta_exports])
            elif platform.system() == "Darwin":
                subprocess.run(["open", pasta_exports])
            else:
                subprocess.run(["xdg-open", pasta_exports])
        except Exception as e:
            print(f"Erro ao abrir pasta: {str(e)}")

    @pyqtSlot()
    def salvar_pesos_automatico(self):
        try:
            for i, tipo in enumerate(self._nomes):
                if i < len(self._valores):
                    peso = self._valores[i]
                    self.mainwindow.gerenciador_pesos.pesos[tipo.strip().upper()] = peso
            
            self.mainwindow.gerenciador_pesos.salvar_pesos()
        except Exception as e:
            print(f"Erro ao salvar pesos automaticamente: {str(e)}")

    @pyqtSlot(int, float)
    def atualizarPeso(self, index, novo_peso):
        try:
            novo_peso = max(1.0, min(10.0, novo_peso))
            
            if 0 <= index < len(self._valores):
                self._valores[index] = int(novo_peso)
                
            self.valoresChanged.emit()
            self.pesosModelChanged.emit(self.pesosModel)
        except Exception as e:
            print(f"Erro ao atualizar peso: {str(e)}")

    @pyqtSlot()
    def salvarPesos(self):
        for i, tipo in enumerate(self._nomes):
            if i < len(self._valores):
                peso = self._valores[i]
                self.mainwindow.gerenciador_pesos.pesos[tipo.strip().upper()] = peso
        
        self.mainwindow.gerenciador_pesos.salvar_pesos()
        print("Pesos salvos com sucesso")

    @pyqtSlot()
    def aplicarPesos(self):
        try:
            for i, tipo in enumerate(self._nomes):
                if i < len(self._valores):
                    peso = self._valores[i]
                    self.mainwindow.gerenciador_pesos.pesos[tipo.strip().upper()] = peso
            
            self.mainwindow.gerenciador_pesos.salvar_pesos()
            print("Pesos aplicados com sucesso")
        except Exception as e:
            print(f"Erro ao aplicar pesos: {str(e)}")

    @pyqtSlot()
    def carregarPesos(self):
        self.mainwindow.gerenciador_pesos.carregar_pesos()
        self.atualizar_tabela_pesos()

    @pyqtSlot()
    def atualizar_dados_completos(self):
        self.atualizar_arquivos_carregados()
        self.atualizar_grafico()
        self.atualizar_kpis()
        self.atualizar_tabela_pesos()

    @pyqtSlot()
    def restaurarPesosPadrao(self):
        tipos_atuais = self._coletar_tipos_brutos()
        self.mainwindow.gerenciador_pesos.restaurar_pesos_padrao(tipos_atuais)
        self.atualizar_tabela_pesos()
        self.atualizar_grafico()
        self.atualizar_kpis()

    @pyqtSlot()
    def testar_todas_funcionalidades(self):
        try:
            print("=== INICIANDO TESTES ===")
            
            print("1. Testando carregamento de arquivos...")
            self.atualizar_arquivos_carregados()
            print(f"   ✓ Meses disponíveis: {len(self.mesesDisponiveis)}")
            
            print("2. Testando KPIs...")
            self.atualizar_kpis()
            kpis = self.kpis()
            print(f"   ✓ KPIs: {kpis}")
            
            print("3. Testando gráfico...")
            self.atualizar_grafico()
            print(f"   ✓ Nomes: {len(self.nomes)}")
            print(f"   ✓ Valores: {len(self.valores)}")
            print(f"   ✓ Max axis: {self.maxAxisValue}")
            
            print("4. Testando pesos...")
            self.atualizar_tabela_pesos()
            print(f"   ✓ Pesos model: {len(self.pesosModel)}")
            
            print("5. Testando tabela semana...")
            self.gerarTabelaSemana()
            print(f"   ✓ Tabela semana: {len(self.tabelaSemana)}")
            
            print("6. Testando ranking...")
            self.atualizar_ranking_model()
            print(f"   ✓ Ranking data: {len(self.rankingData)}")
            
            print("7. Testando filtros...")
            original_count = len(self.tabelaSemana)
            self.filtrarTabelaSemana("test")
            filtered_count = len(self.tabelaSemana)
            self.filtrarTabelaSemana("")
            print(f"   ✓ Filtro funcionando: {original_count} -> {filtered_count}")
            
            print("8. Testando ordenação...")
            self.ordenarTabelaSemana("usuario", True)
            print(f"   ✓ Sort column: {self.sortColumn}")
            print(f"   ✓ Sort ascending: {self.sortAscending}")
            
            print("=== TODOS OS TESTES CONCLUÍDOS ===")
            
        except Exception as e:
            print(f"❌ ERRO NO TESTE: {str(e)}")
            traceback.print_exc()

class KPIs(QObject):
    kpisChanged = pyqtSignal()
    def __init__(self):
        super().__init__()
        self._minutas = "0"
        self._media = "0"
        self._dia = "--"
        self._top3 = "--"

    @pyqtProperty(str, notify=kpisChanged)
    def minutas(self): 
        return self._minutas

    @pyqtProperty(str, notify=kpisChanged)
    def media(self): 
        return self._media

    @pyqtProperty(str, notify=kpisChanged)
    def dia(self): 
        return self._dia

    @pyqtProperty(str, notify=kpisChanged)
    def top3(self): 
        return self._top3

    @pyqtSlot(str, str, str, str)
    def atualizar_kpis(self, minutas, media, dia, top3):
        self._minutas = minutas
        self._media = media
        self._dia = dia
        self._top3 = top3
        self.kpisChanged.emit()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Main Window")
        self.gerenciador_pesos = GerenciadorPesosAgendamento()

    def closeEvent(self, event):
        try:
            if hasattr(self, 'gerenciador_pesos'):
                self.gerenciador_pesos.salvar_pesos()
        except Exception as e:
            print(f"Erro ao salvar pesos ao fechar: {str(e)}")
        finally:
            super().closeEvent(event)


    def chave_mes_ano(self, mes_ano):
        try:
            mes, ano = mes_ano.split("/")
            meses = [
                "jan", "fev", "mar", "abr", "mai", "jun",
                "jul", "ago", "set", "out", "nov", "dez"
            ]
            return int(ano), meses.index(mes.lower())
        except Exception:
            return (0, 0)

    def extrair_mes_ano_do_arquivo(self, file_path):
        try:
            df = self.carregar_planilha(file_path)
            if "Data criação" in df.columns and not df.empty:

                datas = pd.to_datetime(df["Data criação"], errors="coerce", dayfirst=True)
                datas = datas.dropna()
                if not datas.empty:

                    mes_anos = datas.dt.strftime("%m/%Y")
                    mes_ano_mais_frequente = mes_anos.mode()[0]
                    mes, ano = mes_ano_mais_frequente.split("/")

                    MESES_PT = [
                        "jan", "fev", "mar", "abr", "mai", "jun",
                        "jul", "ago", "set", "out", "nov", "dez"
                    ]
                    mes_pt = MESES_PT[int(mes) - 1]
                    return f"{mes_pt}/{ano}"
        except Exception as e:
            print(f"Erro ao extrair mês/ano: {e}")

        return os.path.basename(file_path).replace(".xlsx", "")

    def importar_arquivo_excel(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Importar arquivos Excel",
            "",
            "Todos Arquivos (*);;Excel Files (*.xlsx *.xls)",
        )
        if not paths:
            return
        pasta = "dados_mensais"
        os.makedirs(pasta, exist_ok=True)
        arquivos_importados = 0
        for path in paths:
            try:
                nome_destino = os.path.basename(path)
                destino = os.path.join(pasta, nome_destino)
    
                if path.lower().endswith(".xls"):
                    x2x = XLS2XLSX(path)
                    novo_destino = os.path.splitext(destino)[0] + ".xlsx"
                    x2x.to_xlsx(novo_destino)
                    arquivos_importados += 1
                else:
                    if not os.path.exists(destino):
                        shutil.copy2(path, destino)
                        arquivos_importados += 1
            except Exception as e:
                QMessageBox.warning(
                    self,
                    "Erro",
                    f"Erro ao importar o arquivo '{path}': {str(e)}",
                )
        if arquivos_importados:
            QMessageBox.information(
                self,
                "Importação",
                f"{arquivos_importados} arquivo(s) importado(s) com sucesso!",
            )
        else:
            QMessageBox.warning(
                self,
                "Aviso",
                "Nenhum arquivo novo foi importado (todos já existem na pasta de dados).",
            )
        self.carregar_dados_mensais()
        self.atualizar_listbox_meses()
        self.atualizar_filtros_grafico()
        if hasattr(self, "backend_qml") and self.backend_qml is not None:
            self.backend_qml.atualizar_arquivos_carregados()
            self.backend_qml.atualizar_grafico()
            self.backend_qml.atualizar_kpis()
            self.backend_qml.atualizar_tabela_pesos()
            self.backend_qml.force_chart_update()

    def carregar_dados_mensais(self):
        pasta = "dados_mensais"
        arquivos = (
            [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")]
            if os.path.exists(pasta)
            else []
        )
        dfs = []
        for arq in arquivos:
            file_path = os.path.join(pasta, arq)
            df = self.carregar_planilha(file_path)
            if "Usuário" in df.columns:
                df["Usuário"] = df["Usuário"].astype(str).str.strip().str.upper()
            df = df.loc[:, ~df.columns.duplicated()]
            if not df.empty:
                dfs.append(df)
        self.df = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    def carregar_planilha(self, file_path):
        try:
            if file_path.lower().endswith(".xls"):
                x2x = XLS2XLSX(file_path)
                file_path = x2x.to_xlsx()
            header = pd.read_excel(file_path, nrows=0)
            if "peso" in header.columns:
                df = pd.read_excel(file_path, usecols="A:H")
            else:
                df = pd.read_excel(file_path, skiprows=1, usecols="A:G")
                if "Agendamento" in df.columns:
                    tipo_limpo = df["Agendamento"].apply(
                        lambda x: re.sub(r"\s*\(.*?\)", "", str(x)).strip()
                    )
                    df["peso"] = tipo_limpo.apply(self.gerenciador_pesos.obter_peso)
                else:
                    df["peso"] = 1.0
            if "Usuário" in df.columns:
                df["Usuário"] = df["Usuário"].astype(str).str.strip().str.upper()
            df = df.loc[:, ~df.columns.duplicated()]
            for col in df.columns:
                if isinstance(col, str) and col.lower().startswith("usu") and col != "Usuário":
                    df.rename(columns={col: "Usuário"}, inplace=True)
            
            if "Usuário" in df.columns:
                df = df[~df["Usuário"].isin(["ANGELOBRASIL", "SECAUTOLOC"])]
            
            return df
        except Exception as e:
            print(f"Erro ao carregar planilha {file_path}: {str(e)}")
            return pd.DataFrame()


    def atualizar_listbox_meses(self):
        pasta = "dados_mensais"
        if not hasattr(self, 'listbox_meses'):
            self.listbox_meses = QListWidget()
            self.setCentralWidget(self.listbox_meses)
        arquivos = (
            [arq for arq in os.listdir(pasta) if arq.endswith(".xlsx")]
            if os.path.exists(pasta)
            else []
        )
        opcoes_amigaveis = []
        mapa_arquivo_meses = {}
        for arq in arquivos:
            file_path = os.path.join(pasta, arq)
            mes_ano = self.extrair_mes_ano_do_arquivo(file_path)
            if mes_ano and "/" in mes_ano:
                opcoes_amigaveis.append(mes_ano)
                mapa_arquivo_meses[mes_ano] = arq
        if hasattr(self, 'listbox_meses'):
            self.listbox_meses.clear()
            for mes in sorted(opcoes_amigaveis, key=self.chave_mes_ano):
                self.listbox_meses.addItem(mes)
        else:
            print("listbox_meses is not defined.")
        self.mapa_arquivo_meses = mapa_arquivo_meses

    def atualizar_filtros_grafico(self):
        if getattr(self, 'df', None) is None or self.df.empty:
            for lb in [
                getattr(self, 'filtro_usuario', None),
                getattr(self, 'filtro_mes', None),
                getattr(self, 'filtro_tipo', None),
                getattr(self, 'filtro_agendamento', None),
            ]:
                if lb:
                    lb.clear()
            return

class GerenciadorPesosAgendamento:
    def __init__(self, arquivo_pesos=None):
        if arquivo_pesos is None:
            pasta_destino = os.path.join(os.path.expanduser("~"), "Analyzer-Dev-APROD", "salvos")
            os.makedirs(pasta_destino, exist_ok=True)
            arquivo_pesos = os.path.join(pasta_destino, "pesos.json")
        
        self.arquivo_pesos = arquivo_pesos
        self.pesos = {}
        self.carregar_pesos()

    def restaurar_pesos_padrao(self, tipos_atuais: list):
        for tipo in tipos_atuais:
            self.pesos[tipo.strip().upper()] = 1
        self.salvar_pesos()
        print(f"Pesos restaurados para 1: {len(tipos_atuais)} tipos")

    def carregar_pesos(self):
        try:
            with open(self.arquivo_pesos, 'r', encoding='utf-8') as f:
                dados = json.load(f)
                self.pesos = dict(dados)
        except (FileNotFoundError, json.JSONDecodeError):
            self.pesos = {}

    def salvar_pesos(self):
        try:
            with open(self.arquivo_pesos, 'w', encoding='utf-8') as f:
                json.dump(dict(self.pesos), f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"ERRO CRÍTICO AO SALVAR PESOS: {str(e)}")
            raise

    def _valor_padrao(self):
        return 1.0

    def obter_peso(self, tipo_agendamento):

        tipo_normalizado = str(tipo_agendamento).strip().upper()
        peso = self.pesos.get(tipo_normalizado)
        

        if peso is None:

            tipo_sem_parenteses = re.sub(r'\s*\(.*?\)', '', tipo_normalizado).strip()
            peso = self.pesos.get(tipo_sem_parenteses)
            

            if peso is None:
                peso = 1.0
                
        return float(peso)


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding='utf-8')

        filtered_args = [arg for arg in sys.argv if not arg.startswith('ljsdebugger=')]
        app = QApplication(filtered_args)
        app.setApplicationName("KPI App")
        engine = QQmlApplicationEngine()


        os.makedirs("dados_mensais", exist_ok=True)

        mainwindow = MainWindow()
        backend = Backend(mainwindow)
        kpis = KPIs()
        mainwindow.kpis_qml = kpis
        mainwindow.backend_qml = backend


        engine.rootContext().setContextProperty("mainwindow", mainwindow)
        engine.rootContext().setContextProperty("backend", backend)
        engine.rootContext().setContextProperty("kpis", kpis)


        backend.atualizar_arquivos_carregados()

        backend.atualizar_grafico()
        backend.atualizar_kpis()
        print("Gerando tabela da semana na inicialização...")
        backend.gerarTabelaSemana() 

        qml_file = os.path.join(os.path.dirname(__file__), "MainWindow.qml")
        engine.load(qml_file)
        
        if not engine.rootObjects():
            print("ERROR: Falha ao carregar arquivo QML!")
            sys.exit(1)
            
        sys.exit(app.exec())
    except Exception as e:
        print(f"ERRO DE INICIALIZAÇÃO: {str(e)}")
        import traceback
        traceback.print_exc()
